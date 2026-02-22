import os
import re
import difflib
import openpyxl
from copy import copy
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import pdfplumber
import customtkinter as ctk
from tkinter import filedialog, messagebox
import threading

# --- Business Logic Classes ---

class ProcurementApp:
    def __init__(self):
        self.FILA_INI = 15
        self.PRV_DATA_COLS = {
            "P1": (15, 16, 17),
            "P2": (18, 19, 20),
            "P3": (21, 22, 23),
        }

    def clean_text(self, text):
        return re.sub(r'\s+', ' ', str(text)).strip()

    def match_description(self, desc1, desc2):
        if not desc1 or not desc2:
            return 0
        return difflib.SequenceMatcher(None, str(desc1).upper(), str(desc2).upper()).ratio()

    # Phase 1: Generate Comparison Matrix
    def generate_matrix(self, req_path, bom_path, template_path, output_path):
        try:
            req_wb = openpyxl.load_workbook(req_path, data_only=True)
            bom_wb = openpyxl.load_workbook(bom_path, data_only=True)
            tpl_wb = openpyxl.load_workbook(template_path)
            
            def get_best_sheet(wb, hint=""):
                # Try to find by name hint
                for sheet in wb.worksheets:
                    if hint.lower() in sheet.title.lower():
                        return sheet
                # Otherwise find first sheet with data
                for sheet in wb.worksheets:
                    if sheet.max_row > 2:
                        return sheet
                return wb.active

            req_ws = get_best_sheet(req_wb, "req")
            bom_ws = get_best_sheet(bom_wb, "bom")
            tpl_ws = get_best_sheet(tpl_wb, "comp")
            
            print(f"Phase 1: Using sheets - Req: {req_ws.title}, BOM: {bom_ws.title}, Tpl: {tpl_ws.title}")

            def get_val(ws, cell):
                c = ws[cell]
                val = ""
                if c.value is None: val = ""
                else: val = str(c.value).strip()
                print(f"Phase 1: Extracted {cell} -> '{val}'")
                return val

            # Extract Requisition Metadata
            enc = {
                "obra": get_val(req_ws, "C10"),
                "direccion": get_val(req_ws, "C11"),
                "area": get_val(req_ws, "C12"),
                "solicita": get_val(req_ws, "C13"),
                "descripcion": get_val(req_ws, "C14"),
                "num_req": get_val(req_ws, "F10"),
                "fecha": get_val(req_ws, "F12"),
                "ubicacion": get_val(req_ws, "F13"),
                "especialidad": get_val(req_ws, "F14"),
                "clave_plano": get_val(req_ws, "F15"),
            }

            # Load BOM Catalog
            cat = {}
            for row in bom_ws.iter_rows(values_only=True):
                clave = row[0]
                if not clave or str(clave).strip().upper() == "CLAVE":
                    continue
                clave_str = str(clave).strip()
                
                # SMART COLUMN DETECTION:
                desc = ""
                unit = ""
                cost = 0

                # 1. Find Description: The longest string in the first 6 columns
                max_len = 0
                for i in range(1, 6):
                    val = str(row[i]).strip() if len(row) > i and row[i] else ""
                    if len(val) > max_len:
                        max_len = len(val)
                        desc = val

                # 2. Find Unit: A short string that matches common unit patterns
                unit_patterns = ["PZA", "ML", "M", "KG", "L", "TON", "M2", "M3", "SRV", "LOTE", "PJE"]
                for i in range(1, 6):
                    val = str(row[i]).strip().upper() if len(row) > i and row[i] else ""
                    if val in unit_patterns:
                        unit = val
                        break
                
                # Fallback for unit
                if not unit:
                    unit = str(row[2]).strip() if len(row) > 2 and row[2] else (str(row[16]).strip() if len(row) > 16 and row[16] else "")

                # 3. Find Cost
                try:
                    cost = float(row[23]) if len(row) > 23 and row[23] is not None else (float(row[4]) if len(row) > 4 and row[4] is not None else 0)
                except:
                    cost = 0
                
                cat[clave_str] = {"descripcion": desc, "unidad": unit, "costo": cost}

            # Fill Template Headers
            tpl_ws["C5"] = enc["obra"]
            tpl_ws["G5"] = enc["num_req"]
            tpl_ws["C6"] = "" # Contrato (Unassigned)
            tpl_ws["G6"] = enc["fecha"]
            tpl_ws["C7"] = enc["direccion"]
            tpl_ws["G7"] = enc["ubicacion"]
            tpl_ws["C8"] = enc["area"]
            tpl_ws["G8"] = enc["especialidad"]
            tpl_ws["C9"] = enc["solicita"]
            tpl_ws["G9"] = enc["clave_plano"]
            tpl_ws["C10"] = enc["descripcion"]

            # Populate Materials
            materiales_agregados = 0
            current_row = self.FILA_INI
            
            print(f"Phase 1: Starting row iteration from row 18. Total rows: {req_ws.max_row}")

            for row_idx, row in enumerate(req_ws.iter_rows(min_row=18, values_only=True), start=18):
                # Try Col 2 (B) then Col 1 (A)
                clave = row[1] if len(row) > 1 else row[0]
                if not clave or str(clave).strip() == "" or str(clave).strip().upper() in ["CLAVE", "PARTIDA"]:
                    continue
                
                clave_str = str(clave).strip()
                req_qty = row[4] if len(row) > 4 else (row[3] if len(row) > 3 else 0)
                try:
                    req_qty = float(req_qty)
                except:
                    req_qty = 0
                
                print(f"Phase 1: Found item {clave_str} at row {row_idx}")
                
                info = cat.get(clave_str, {})
                costo = info.get("costo", 0) or 0
                req_desc = str(row[2]).strip() if len(row) > 2 and row[2] else ""
                req_unit = str(row[3]).strip() if len(row) > 3 and row[3] else ""
                
                # Ensure we have a description and unit
                final_desc = info.get("descripcion") or req_desc
                final_unit = info.get("unidad") or req_unit

                # Fill Matrix Row
                tpl_ws.cell(row=current_row, column=2).value = clave_str
                tpl_ws.cell(row=current_row, column=3).value = final_desc
                tpl_ws.cell(row=current_row, column=4).value = final_unit
                tpl_ws.cell(row=current_row, column=5).value = req_qty
                tpl_ws.cell(row=current_row, column=6).value = costo
                
                # Formulas and constants
                tpl_ws.cell(row=current_row, column=7).value = f"=E{current_row}*F{current_row}"
                tpl_ws.cell(row=current_row, column=8).value = -0.05
                tpl_ws.cell(row=current_row, column=9).value = f"=G{current_row}*H{current_row}"
                tpl_ws.cell(row=current_row, column=10).value = f"=G{current_row}+I{current_row}"
                
                # Initialize Winner Columns (X, Y, Z -> 24, 25, 26) with 0
                tpl_ws.cell(row=current_row, column=24).value = 0
                tpl_ws.cell(row=current_row, column=25).value = 0
                tpl_ws.cell(row=current_row, column=26).value = 0
                
                current_row += 1
                materiales_agregados += 1

            tpl_wb.save(output_path)
            return True, f"Matrix generated with {materiales_agregados} items."
        except Exception as e:
            return False, f"Error in Phase 1: {str(e)}"

    # Phase 2: Extract PDF and Populate Matrix
    def extract_and_populate(self, matrix_path, pdf_paths, bom_path):
        try:
            wb = openpyxl.load_workbook(matrix_path)
            ws = wb.active
            
            # Load BOM for matching
            bom_wb = openpyxl.load_workbook(bom_path, data_only=True)
            bom_ws = bom_wb.active
            cat = {}
            for row in bom_ws.iter_rows(values_only=True):
                clave = row[0]
                if not clave or str(clave).strip().upper() == "CLAVE": continue
                cat[str(clave).strip()] = {"descripcion": str(row[2]).strip() if len(row) > 2 and row[2] else ""}

            # Match and Update
            clave_a_fila = {}
            for r in range(self.FILA_INI, ws.max_row + 1):
                clave = ws.cell(r, 2).value
                if clave and "-" in str(clave):
                    clave_a_fila[str(clave).strip()] = r

            # Process each PDF into a separate slot (P1, P2, P3)
            # Create or get Suppliers_DB sheet
            if "Suppliers_DB" not in wb.sheetnames:
                db_ws = wb.create_sheet("Suppliers_DB")
                db_ws.append(['Supplier Slot', 'Name', 'Contact', 'Email', 'Delivery Time (Avg)'])
            else:
                db_ws = wb["Suppliers_DB"]

            for i, pdf_path in enumerate(pdf_paths[:3]):
                try:
                    if not pdf_path.lower().endswith('.pdf'):
                        print(f"Skipping non-PDF file: {pdf_path}")
                        continue

                    slot_key = f"P{i+1}"
                    print(f"Phase 2: Processing PDF {i+1}: {os.path.basename(pdf_path)}")
                    col_cost, col_imp, col_te = self.PRV_DATA_COLS[slot_key]
                    meta_col = {0: 3, 1: 9, 2: 15}[i] # Col C, I, O in Row 91

                    all_pdf_items = []
                    supplier_name = os.path.basename(pdf_path).replace('.pdf', '')
                    
                    with pdfplumber.open(pdf_path) as pdf_doc:
                        full_text = ""
                        tables = []
                        for page in pdf_doc.pages:
                            full_text += (page.extract_text() or "") + "\n"
                            tables.extend(page.extract_table() or [])
                        
                        # Basic meta extraction
                        supplier_name = self.extract_regex(full_text, r'Supplier:\s*(.*)') or \
                                        self.extract_regex(full_text, r'Proveedor:\s*(.*)') or supplier_name
                        
                        # Extract items from tables
                        for table in tables:
                            if not table or len(table) < 2: continue
                            for row in table[1:]:
                                if not row: continue
                                desc = str(row[1]).strip() if len(row) > 1 else ""
                                price = self.parse_price(row[2]) if len(row) > 2 else None
                                if desc and price:
                                    all_pdf_items.append({"desc": desc, "price": price})

                    # Update Matrix with this PDF's data
                    for item in all_pdf_items:
                        best_clave = None
                        best_score = 0
                        for clave, info in cat.items():
                            score = self.match_description(item["desc"], info["descripcion"])
                            if score > 0.42 and score > best_score:
                                best_clave = clave
                                best_score = score
                        
                        if best_clave and best_clave in clave_a_fila:
                            r = clave_a_fila[best_clave]
                            ws.cell(row=r, column=col_cost).value = item["price"]
                            ws.cell(row=r, column=col_te).value = "8-12 Days"

                    # Update Supplier Meta (Row 91)
                    ws.cell(row=91, column=meta_col).value = supplier_name
                    # Update DB Sheet
                    db_ws.append([slot_key, supplier_name, "Extracted from PDF", "N/A", "See T.E. Column"])
                except Exception as pdf_e:
                    print(f"Error processing PDF {pdf_path}: {pdf_e}")

            # Add Winner Selection Headers
            ws.cell(row=14, column=24).value = "CHOOSE P1"
            ws.cell(row=14, column=25).value = "CHOOSE P2"
            ws.cell(row=14, column=26).value = "CHOOSE P3"

            # Add Summary Row below the last item row to avoid overwriting data
            max_item_row = max(clave_a_fila.values()) if clave_a_fila else 80
            summary_row = max_item_row + 1
            ws.cell(row=summary_row, column=2).value = "TOTAL ITEMS SELECTED"
            ws.cell(row=summary_row, column=24).value = f'=COUNTIF(X15:X{max_item_row},1)+COUNTIF(X15:X{max_item_row},"X")+COUNTIF(X15:X{max_item_row},"x")+COUNTIF(X15:X{max_item_row},"✓")'
            ws.cell(row=summary_row, column=25).value = f'=COUNTIF(Y15:Y{max_item_row},1)+COUNTIF(Y15:Y{max_item_row},"X")+COUNTIF(Y15:Y{max_item_row},"x")+COUNTIF(Y15:Y{max_item_row},"✓")'
            ws.cell(row=summary_row, column=26).value = f'=COUNTIF(Z15:Z{max_item_row},1)+COUNTIF(Z15:Z{max_item_row},"X")+COUNTIF(Z15:Z{max_item_row},"x")+COUNTIF(Z15:Z{max_item_row},"✓")'

            wb.save(matrix_path)
            return True, f"Matrix updated with {len(pdf_paths)} PDF(s)."
        except Exception as e:
            return False, f"Error in Phase 2: {str(e)}"

    def extract_regex(self, text, pattern):
        match = re.search(pattern, text, re.IGNORECASE)
        return match.group(1).strip() if match else None

    def parse_price(self, text):
        if not text: return None
        clean = str(text).replace('$', '').replace(',', '').strip()
        try:
            return float(clean)
        except:
            return None

    # Phase 3: Generate POs
    def generate_pos(self, matrix_path, po_template_path, output_dir):
        try:
            wb = openpyxl.load_workbook(matrix_path, data_only=True)
            sheet = wb.active
            
            def get_val(ws, cell):
                c = ws[cell]
                if c.value is None:
                    return ""
                return str(c.value).strip()

            def get_first_val(ws, *cells):
                for cell in cells:
                    value = get_val(ws, cell)
                    if value:
                        return value
                return ""

            def to_float(value):
                if value is None:
                    return 0.0
                try:
                    return float(str(value).replace(",", "").replace("$", "").strip())
                except Exception:
                    return 0.0

            # Extract Metadata from Matrix
            enc = {
                "obra": get_first_val(sheet, "C5"),
                "num_req": get_first_val(sheet, "G5", "F5"),
                "contrato": get_val(sheet, "C6"),
                "fecha": get_first_val(sheet, "G6", "F6"),
                "direccion": get_val(sheet, "C7"),
                "ubicacion": get_first_val(sheet, "G7", "F7"),
                "solicita": get_val(sheet, "C9"),
            }

            # Find winners (Checkboxes in col X, Y, Z -> 24, 25, 26)
            winners = {0: [], 1: [], 2: []} # Supplier index -> list of items
            total_winners = 0
            
            for row_idx in range(self.FILA_INI, sheet.max_row + 1):
                item_code = sheet.cell(row=row_idx, column=2).value
                if not item_code:
                    continue

                item_code_str = str(item_code).strip()
                upper_code = item_code_str.upper()
                if upper_code in ["CLAVE", "PARTIDA"]:
                    continue
                if "TOTAL ITEMS SELECTED" in upper_code:
                    break
                
                for s_idx in range(3):
                    check_col = 24 + s_idx
                    val = str(sheet.cell(row=row_idx, column=check_col).value).strip().upper()
                    if val in ['✓', 'X', '1', 'TRUE', 'SI', 'SÍ']:
                        qty_value = to_float(sheet.cell(row=row_idx, column=5).value)
                        price_col = 15 + (s_idx * 3)  # O, R, U
                        price_value = to_float(sheet.cell(row=row_idx, column=price_col).value)
                        total_col = 16 + (s_idx * 3)  # P, S, V
                        line_total = to_float(sheet.cell(row=row_idx, column=total_col).value) or (qty_value * price_value)

                        winners[s_idx].append({
                            'code': item_code_str,
                            'desc': sheet.cell(row=row_idx, column=3).value,
                            'unit': sheet.cell(row=row_idx, column=4).value,
                            'qty': qty_value,
                            'price': price_value,
                            'line_total': line_total
                        })
                        total_winners += 1

            if total_winners == 0:
                return False, "No winners selected. Please mark winners with a '1', 'X', or '✓' in columns X, Y, or Z."

            generated_count = 0
            for s_idx, items in winners.items():
                if not items: continue
                
                # Load supplier info (Row 91, Col C=3, I=9, O=15)
                col_s = 3 + (s_idx * 6)
                supplier_name = str(sheet.cell(row=91, column=col_s).value or f"Supplier_{s_idx+1}").strip()
                sanitized_name = "".join([c if c.isalnum() or c in (' ', '-', '_') else '-' for c in supplier_name])
                
                po_wb = openpyxl.load_workbook(po_template_path)
                po_ws = po_wb.active
                
                # Fill PO Headers
                po_ws['D9'] = supplier_name
                po_ws['M9'] = enc["obra"]
                po_ws['M10'] = enc["direccion"]
                po_ws['M11'] = enc["ubicacion"]
                po_ws['M15'] = enc["num_req"]
                po_ws['N5'] = enc["fecha"]
                po_ws['N6'] = enc["solicita"]
                
                # Expand items area if selected rows are greater than template capacity
                frame_end = 34
                base_item_start = 19
                base_item_end = 34
                base_totals_row = 35
                base_capacity = base_item_end - base_item_start + 1
                extra_rows = max(0, len(items) - base_capacity)

                if extra_rows > 0:
                    # Insert inside the PO frame so the frame grows from B19:B34.
                    po_ws.insert_rows(frame_end, amount=extra_rows)
                    # Copy regular item-row style into the inserted rows (use row 33 as source).
                    style_source_row = base_item_end - 1
                    for r_new in range(frame_end, frame_end + extra_rows):
                        for c_new in range(1, po_ws.max_column + 1):
                            src_cell = po_ws.cell(row=style_source_row, column=c_new)
                            dst_cell = po_ws.cell(row=r_new, column=c_new)
                            if src_cell.has_style:
                                dst_cell._style = copy(src_cell._style)

                totals_row = base_totals_row + extra_rows

                # Clear items area
                for r_clear in range(base_item_start, totals_row):
                    for c_clear in [2, 5, 14, 13, 15, 16]:
                        po_ws.cell(row=r_clear, column=c_clear).value = None
                
                start_row = base_item_start

                for i, item in enumerate(items):
                    r = start_row + i

                    po_ws.cell(row=r, column=2).value = item['code']
                    po_ws.cell(row=r, column=5).value = item['desc']
                    po_ws.cell(row=r, column=14).value = item['qty']
                    po_ws.cell(row=r, column=13).value = item['unit']
                    po_ws.cell(row=r, column=15).value = item['price']
                    po_ws.cell(row=r, column=16).value = f"=N{r}*O{r}"
                
                # Fill dynamic totals row
                last_item_row = start_row + len(items) - 1
                po_ws.cell(row=totals_row, column=14).value = f"=SUM(N{start_row}:N{last_item_row})"
                po_ws.cell(row=totals_row, column=16).value = f"=SUM(P{start_row}:P{last_item_row})"
                
                out_name = f"PO_{enc['num_req']}_P{s_idx+1}_{sanitized_name}.xlsx"
                po_wb.save(os.path.join(output_dir, out_name))
                generated_count += 1

            return True, f"Generated {generated_count} POs."
        except Exception as e:
            return False, f"Error in Phase 3: {str(e)}"

# --- GUI Implementation ---

class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("ProcureFlow AI - Automation Suite")
        self.geometry("900x600")
        ctk.set_appearance_mode("dark")
        ctk.set_default_color_theme("blue")
        
        self.engine = ProcurementApp()
        
        # Sidebar
        self.sidebar = ctk.CTkFrame(self, width=200, corner_radius=0)
        self.sidebar.pack(side="left", fill="y")
        
        self.logo = ctk.CTkLabel(self.sidebar, text="ProcureFlow", font=ctk.CTkFont(size=20, weight="bold"))
        self.logo.pack(pady=20)
        
        self.tab_btns = []
        for i, name in enumerate(["Phase 1: Matrix", "Phase 2: Extract", "Phase 3: POs"]):
            btn = ctk.CTkButton(self.sidebar, text=name, command=lambda x=i: self.show_tab(x))
            btn.pack(pady=10, padx=20)
            self.tab_btns.append(btn)
            
        # Main Content
        self.container = ctk.CTkFrame(self)
        self.container.pack(side="right", fill="both", expand=True, padx=20, pady=20)
        
        self.tabs = []
        self.create_phase1()
        self.create_phase2()
        self.create_phase3()
        
        self.show_tab(0)

    def show_tab(self, index):
        for tab in self.tabs: tab.pack_forget()
        self.tabs[index].pack(fill="both", expand=True)

    def create_phase1(self):
        tab = ctk.CTkFrame(self.container)
        self.tabs.append(tab)
        
        ctk.CTkLabel(tab, text="Generate Comparison Matrix", font=("Arial", 18, "bold")).pack(pady=10)
        
        self.req_file = self.create_file_row(tab, "Requisition (XLSX)")
        self.bom_file = self.create_file_row(tab, "BOM (XLSX)")
        self.tpl_file = self.create_file_row(tab, "Matrix Template (XLSX)")
        
        ctk.CTkButton(tab, text="Generate Draft Matrix", command=self.run_phase1).pack(pady=30)

    def create_phase2(self):
        tab = ctk.CTkFrame(self.container)
        self.tabs.append(tab)
        
        ctk.CTkLabel(tab, text="Extract Supplier Quotes", font=("Arial", 18, "bold")).pack(pady=10)
        
        self.matrix_file = self.create_file_row(tab, "Draft Matrix (XLSX)")
        self.bom_file_p2 = self.create_file_row(tab, "BOM (XLSX)")
        
        self.pdf_list = []
        ctk.CTkButton(tab, text="Add Supplier PDF", command=self.add_pdf).pack(pady=5)
        self.pdf_label = ctk.CTkLabel(tab, text="0 PDFs selected")
        self.pdf_label.pack()
        
        ctk.CTkButton(tab, text="Extract & Populate", command=self.run_phase2).pack(pady=30)

    def create_phase3(self):
        tab = ctk.CTkFrame(self.container)
        self.tabs.append(tab)
        
        ctk.CTkLabel(tab, text="Generate Purchase Orders", font=("Arial", 18, "bold")).pack(pady=10)
        
        self.final_matrix = self.create_file_row(tab, "Final Matrix (XLSX)")
        self.po_tpl = self.create_file_row(tab, "PO Template (XLSX)")
        self.out_dir = self.create_file_row(tab, "Output Directory", is_dir=True)
        
        ctk.CTkButton(tab, text="Generate All POs", command=self.run_phase3).pack(pady=30)

    def create_file_row(self, parent, label, is_dir=False):
        frame = ctk.CTkFrame(parent, fg_color="transparent")
        frame.pack(fill="x", padx=20, pady=5)
        ctk.CTkLabel(frame, text=label, width=150, anchor="w").pack(side="left")
        entry = ctk.CTkEntry(frame)
        entry.pack(side="left", fill="x", expand=True, padx=10)
        btn = ctk.CTkButton(frame, text="Browse", width=80, 
                            command=lambda: self.browse(entry, is_dir))
        btn.pack(side="right")
        return entry

    def browse(self, entry, is_dir):
        path = filedialog.askdirectory() if is_dir else filedialog.askopenfilename()
        if path:
            entry.delete(0, "end")
            entry.insert(0, path)

    def add_pdf(self):
        files = filedialog.askopenfilenames(filetypes=[("PDF files", "*.pdf")])
        if files:
            self.pdf_list.extend(files)
            self.pdf_label.configure(text=f"{len(self.pdf_list)} PDFs selected")

    def run_phase1(self):
        def task():
            success, msg = self.engine.generate_matrix(
                self.req_file.get(), self.bom_file.get(), self.tpl_file.get(), "Draft_Matrix.xlsx"
            )
            messagebox.showinfo("Result", msg)
        threading.Thread(target=task).start()

    def run_phase2(self):
        def task():
            success, msg = self.engine.extract_and_populate(
                self.matrix_file.get(), 
                self.pdf_list, 
                self.bom_file_p2.get()
            )
            messagebox.showinfo("Result", msg)
        threading.Thread(target=task).start()

    def run_phase3(self):
        def task():
            success, msg = self.engine.generate_pos(
                self.final_matrix.get(), self.po_tpl.get(), self.out_dir.get()
            )
            messagebox.showinfo("Result", msg)
        threading.Thread(target=task).start()

if __name__ == "__main__":
    app = App()
    app.mainloop()
